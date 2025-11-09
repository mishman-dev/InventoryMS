from django.shortcuts import render,get_object_or_404
from django.db.models import F, Prefetch
from django.contrib.admin.models import LogEntry
from django.contrib.contenttypes.models import ContentType
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from .models import Item, Category
import random, string
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from .models import Issue, Employee, Project, Item, Supplier, Purchase
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.utils.dateparse import parse_date
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation
import uuid


def issue_list(request):
    issues = Issue.objects.all().order_by('-issue_date', '-id')
    today_issues = issues.filter(issue_date=datetime.today().date()).count()
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        issues = issues.filter(
            Q(utilize_no__icontains=search_query) |
            Q(issued_to__first_name__icontains=search_query) |
            Q(issued_to__last_name__icontains=search_query) |
            Q(item__name__icontains=search_query) |
            Q(project__name__icontains=search_query)
        )
    
    # Filter by project
    project_filter = request.GET.get('project', '')
    if project_filter:
        issues = issues.filter(project_id=project_filter)
    
    # Filter by employee
    employee_filter = request.GET.get('employee', '')
    if employee_filter:
        issues = issues.filter(issued_to_id=employee_filter)
    
    context = {
        'issues': issues,
        'projects': Project.objects.all(),
        'employees': Employee.objects.all(),
        'search_query': search_query,
        'project_filter': project_filter,
        'employee_filter': employee_filter,
        'today_issues':today_issues,
    }
    
    return render(request, 'issues/issue_list.html', context)


def export_issues_excel(request):
    """Export issues to Excel with date range filter"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Filter issues by date range
    issues = Issue.objects.all().select_related('issued_to', 'project', 'item')
    
    if start_date:
        issues = issues.filter(issue_date__gte=start_date)
    if end_date:
        issues = issues.filter(issue_date__lte=end_date)
    
    issues = issues.order_by('-issue_date')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Issues Report"
    
    # Define header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "Issues Report"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add date range info
    if start_date or end_date:
        ws.merge_cells('A2:H2')
        date_info = ws['A2']
        date_range_text = f"Period: {start_date or 'Beginning'} to {end_date or 'Present'}"
        date_info.value = date_range_text
        date_info.alignment = Alignment(horizontal="center")
    
    # Headers
    header_row = 4 if (start_date or end_date) else 3
    headers = ['Utilize No', 'Issue Date', 'Issued To', 'Project', 'Item', 'Quantity', 'Unit', 'Remarks']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add data
    data_start_row = header_row + 1
    for row_num, issue in enumerate(issues, data_start_row):
        ws.cell(row=row_num, column=1, value=issue.utilize_no)
        ws.cell(row=row_num, column=2, value=issue.issue_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=3, value=f"{issue.issued_to.first_name} {issue.issued_to.last_name}")
        ws.cell(row=row_num, column=4, value=issue.project.name if issue.project else 'N/A')
        ws.cell(row=row_num, column=5, value=issue.item.name)
        ws.cell(row=row_num, column=6, value=issue.quantity_issued)
        ws.cell(row=row_num, column=7, value=issue.item.unit if hasattr(issue.item, 'unit') else '')
        ws.cell(row=row_num, column=8, value=issue.remarks or '')
    
    # Adjust column widths
    column_widths = [15, 12, 20, 20, 25, 10, 10, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"issues_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def dashboard(request):
    items = Item.objects.select_related('category').all()
    all_items = Item.objects.all()
    categories = Category.objects.all()
    dashboard_items = Item.objects.all()[:6]
    # Calculate statistics
    total_products = items.count()
    in_stock = items.filter(current_stock__gt=0).count()
    low_stock = items.filter(current_stock__lte=F('reorder_level'), current_stock__gt=0)
    out_of_stock = items.filter(current_stock=0)
    recent_logs = LogEntry.objects.select_related('content_type', 'user').order_by('-action_time')[:5]

    context = {
        'items': items,
        'total_products': total_products,
        'in_stock': in_stock,
        'low_stock': low_stock,
        'out_of_stock': out_of_stock,
        'recent_logs': recent_logs,
        'activities': recent_logs,
        'all_items':all_items,
        'dashboard_items':dashboard_items,
        'categories':categories,
    }
    return render(request, 'inventory/dashboard.html', context)

def item_tab(request):
    items = Item.objects.all().order_by('name')
    items_cat = Item.objects.all().select_related('category_name')
    filter_type = request.GET.get('filter', 'all')
    search_query = request.GET.get('search', '')
    categories = Category.objects.all()
    suppliers = Supplier.objects.all()

    # Filter logic
    if filter_type == 'low_stock':
        items = items.filter(current_stock__lte=F('reorder_level'), current_stock__gt=0)
    elif filter_type == 'out_of_stock':
        items = items.filter(current_stock=0)
    elif filter_type == 'expired_warranty':
        items = items.filter(warrenty__lt=datetime.today().date())

    # Search logic
    if search_query:
        items = items.filter(name__icontains=search_query)

    context = {
        'items_cat': items_cat,
        'filter_type': filter_type,
        'search_query': search_query,
        'items': items,
        'categories':categories,
        'suppliers':suppliers,
    }
    return render(request, 'inventory/items.html', context)

def generate_item_code():
    """Generate unique item code like 'ITM-AX45G9'"""
    while True:
        code = 'CCST-' + ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
        if not Item.objects.filter(item_code=code).exists():
            return code

def category_list(request):
    if request.method == "POST":
        # Add new category
        name = request.POST.get("category_name")
        if name:
            category_id = get_random_string(8).upper()
            Category.objects.create(category_id=category_id, category_name=name)
            messages.success(request, "Category added successfully!")
        return redirect("category_list")

    categories = Category.objects.all().order_by("category_name")
    return render(request, "inventory/category_list.html", {"categories": categories})

def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    category.delete()
    messages.success(request, "Category deleted successfully!")
    return redirect("category_list")

def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        category.category_name = request.POST.get("category_name")
        category.save()
        messages.success(request, "Category updated successfully!")
        return redirect("category_list")
    return render(request, "inventory/category_edit.html", {"category": category})


@csrf_exempt   # <-- prefer to REMOVE this in production and rely on CSRF token from the form
def add_item(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)

    try:
        # READ values
        name = request.POST.get('name', '').strip()
        category_id = request.POST.get('category_name')   # you send category_id from select
        default_unit_of_measure = request.POST.get('default_unit_of_measure', '').strip()
        unit_price_raw = request.POST.get('unit_price', '').strip()
        current_stock_raw = request.POST.get('current_stock', '').strip()
        supplier_id = request.POST.get('supplier_id').strip()
        reorder_level_raw = request.POST.get('reorder_level', '').strip()
        warranty_raw = request.POST.get('warrenty') or request.POST.get('warranty')  # accept both spellings

        # VALIDATION basics
        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'}, status=400)

        # Resolve category (allow blank)
        category_obj = None
        if category_id:
            # You used category.category_id in the select, so lookup by that
            category_obj = Category.objects.filter(category_id=category_id).first()
            if not category_obj:
                return JsonResponse({'success': False, 'error': 'Selected category not found'}, status=400)
        
        supplier_obj =None
        if supplier_id:
            supplier_obj = Supplier.objects.filter(supplier_id=supplier_id).first()
            if not supplier_obj:
                return JsonResponse({'success': False, 'error': 'Selected Supplier not found'}, status=400)

        # Convert numeric fields safely
        try:
            unit_price = Decimal(unit_price_raw) if unit_price_raw != '' else Decimal('0.00')
        except (InvalidOperation, ValueError):
            return JsonResponse({'success': False, 'error': 'Invalid unit price'}, status=400)

        try:
            current_stock = int(current_stock_raw) if current_stock_raw != '' else 0
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid current stock'}, status=400)

        try:
            reorder_level = int(reorder_level_raw) if reorder_level_raw != '' else 0
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid reorder level'}, status=400)

        # Parse warranty date (optional)
        warranty = None
        if warranty_raw:
            warranty = parse_date(warranty_raw)   # returns a date or None
            if warranty is None:
                return JsonResponse({'success': False, 'error': 'Invalid warranty date'}, status=400)

        # Generate item code
        item_code = generate_item_code()   # make sure this function exists and returns a string

        # Create item - pass the category object (or None) to the FK
        item = Item.objects.create(
            item_code=item_code,
            name=name,
            category=category_obj,
            warrenty=warranty,
            default_unit_of_measure=default_unit_of_measure,
            unit_price=unit_price,
            supplier=supplier_obj,
            current_stock=current_stock,
            reorder_level=reorder_level,
        )

        # Prepare response (send category display name for client)
        category_display = category_obj.category_name if category_obj else None

        return JsonResponse({
            'success': True,
            'item': {
                'id': item.id,
                'item_code': item.item_code,
                'name': item.name,
                'category_name': category_display,
                'unit_price': float(item.unit_price),
                'current_stock': item.current_stock,
                'reorder_level': item.reorder_level,
                'warrenty': item.warrenty.strftime('%Y-%m-%d') if item.warrenty else "N/A",
            }
        })

    except Exception as e:
        # Log server-side in console for debugging (optional)
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def generate_utilize_no():
    """Generate a unique utilization number"""
    prefix = "UTZ-"
    random_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
    return f"{prefix}{random_code}"



def add_issue(request):
    if request.method == 'POST':
        try:
            utilize_no = generate_utilize_no()
            issued_to_id = request.POST.get('issued_to')
            project_id = request.POST.get('project')
            remarks = request.POST.get('remarks', '')

            items = request.POST.getlist('items[]')
            quantities = request.POST.getlist('quantities[]')

            # Validate foreign keys
            issued_to = Employee.objects.get(id=issued_to_id)
            project = Project.objects.get(id=project_id) if project_id else None

            # Create the main Issue record
            issue = Issue.objects.create(
                utilize_no=utilize_no,
                issued_to=issued_to,
                project=project,
                remarks=remarks,
                issue_date=timezone.now().date()
            )
            
            # Create individual IssueItems
            for i in range(len(items)):
                item_obj = Item.objects.get(id=items[i])
                qty = int(quantities[i])
                Issue.objects.create(issue=issue, item=item_obj, quantity_issued=qty)

            return JsonResponse({
                'success': True,
                'issue': {
                    'id': issue.id,
                    'utilize_no': issue.utilize_no,
                    'issued_to': issued_to.first_name,
                    'project': project.name if project else "N/A",
                    'issue_date': issue.issue_date.strftime('%Y-%m-%d'),
                    'remarks': issue.remarks or "N/A",
                    'items': [
                        {
                            'item': Item.objects.get(id=items[i]).name,
                            'quantity_issued': quantities[i]
                        } for i in range(len(items))
                    ]
                }
            })

        except (Item.DoesNotExist, Employee.DoesNotExist, Project.DoesNotExist) as e:
            return JsonResponse({'success': False, 'error': str(e)})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})

def supplier_list(request):
    suppliers = Supplier.objects.all()
    query = request.GET.get('q', '')
    if query:
        suppliers = suppliers.filter(
            Q(name__icontains=query) | 
            Q(contact_person__icontains=query) | 
            Q(phone__icontains=query)
        )
    return render(request, 'inventory/supplier_list.html', {'suppliers': suppliers, 'query': query, })

def add_supplier(request):
    if request.method == 'POST':
        supplier = Supplier(
            supplier_id=str(uuid.uuid4()),
            name=request.POST['name'],
            contact_person=request.POST.get('contact_person'),
            phone=request.POST.get('phone'),
            email=request.POST.get('email'),
            address=request.POST.get('address'),
        )
        supplier.save()
        messages.success(request, 'Supplier added successfully!')
        return redirect('supplier_list')
    return redirect('supplier_list')

def edit_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    if request.method == 'POST':
        supplier.name = request.POST['name']
        supplier.contact_person = request.POST.get('contact_person')
        supplier.phone = request.POST.get('phone')
        supplier.email = request.POST.get('email')
        supplier.address = request.POST.get('address')
        supplier.save()
        messages.success(request, 'Supplier updated successfully!')
        return redirect('supplier_list')
    return render(request, 'edit_supplier.html', {'supplier': supplier})

def delete_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    supplier.delete()
    messages.success(request, 'Supplier deleted successfully!')
    return redirect('supplier_list')

def generate_project_id():
    return f"PRJ-{uuid.uuid4().hex[:8].upper()}"

def project_list(request):
    query = request.GET.get('q', '')
    projects = Project.objects.all()

    if query:
        projects = projects.filter(project_name__icontains=query)

    employees = Employee.objects.all()

    if request.method == 'POST':
        project_name = request.POST.get('project_name', '').strip()
        responsible_person_id = request.POST.get('responsible_person')

        if not project_name:
            return JsonResponse({'success': False, 'error': 'Project name is required'}, status=400)

        emp_obj = Employee.objects.filter(id=responsible_person_id).first() if responsible_person_id else None

        Project.objects.create(
            project_id=generate_project_id(),
            project_name=project_name,
            responsible_person=emp_obj
        )

        return redirect('project_list')

    context = {
        'projects': projects,
        'employees': employees,
        'query': query
    }
    return render(request, 'inventory/project_list.html', context)


def project_edit(request, pk):
    project = get_object_or_404(Project, pk=pk)
    employees = Employee.objects.all()

    if request.method == 'POST':
        project.project_name = request.POST.get('project_name', project.project_name)
        responsible_person_id = request.POST.get('responsible_person')
        if responsible_person_id:
            project.responsible_person = Employee.objects.filter(id=responsible_person_id).first()
        project.save()
        return redirect('project_list')

    return render(request, 'inventory/project_edit.html', {'project': project, 'employees': employees})


def project_delete(request, pk):
    project = get_object_or_404(Project, pk=pk)
    project.delete()
    return redirect('project_list')

def get_item_price(request):
    item_id = request.GET.get('item_id')
    try:
        item = Item.objects.get(id=item_id)
        return JsonResponse({'unit_price': float(item.unit_price)})
    except Item.DoesNotExist:
        return JsonResponse({'unit_price': 0})
    

def generate_purchase_no():
    return f"P-{uuid.uuid4().hex[:8].upper()}"

def purchase_list(request):
    purchases = Purchase.objects.select_related('supplier', 'item').order_by('-purchase_date')
    total_qty = sum(p.quantity for p in purchases)
    total_cost = sum(p.total_cost for p in purchases)
    return render(request, 'inventory/purchases.html', {'purchases': purchases,'total_qty': total_qty,'total_cost': total_cost})

def add_purchase(request):
    if request.method == 'POST':
        supplier = Supplier.objects.get(id=request.POST.get('supplier'))

        items = request.POST.getlist('item[]')
        quantities = request.POST.getlist('quantity[]')
        prices = request.POST.getlist('unit_price[]')

        for item_id, qty, price in zip(items, quantities, prices):
            item = Item.objects.get(id=item_id)
            qty = int(qty)
            price = item.unit_price 

            Purchase.objects.create(
                purchase_no=generate_purchase_no(),
                supplier=supplier,
                item=item,
                quantity=qty,
                unit_price=price,
                total_cost=qty * price,
                purchase_date=timezone.now()
            )

        return redirect('/purchases')  # Correct redirect syntax

    suppliers = Supplier.objects.all()
    items = Item.objects.all()
    return render(request, 'inventory/add_purchase.html', {'suppliers': suppliers, 'items': items})

def employee_list(request):
    query = request.GET.get('q', '')
    if query:
        employees = Employee.objects.filter(
            Q(first_name__icontains=query) | Q(last_name__icontains=query)
        )
    else:
        employees = Employee.objects.all()

    return render(request, 'inventory/employee_list.html', {'employees': employees, 'query': query})

def add_employee(request):
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        phone_number = request.POST.get('phone_number')
        department = request.POST.get('department')
        position = request.POST.get('position')

        # generate a unique employee ID
        employee_id = f"EMP-{uuid.uuid4().hex[:6].upper()}"

        Employee.objects.create(
            employee_id=employee_id,
            first_name=first_name,
            last_name=last_name,
            phone_number=phone_number,
            department=department,
            position=position
        )
        return redirect('employee_list')
    return JsonResponse({'error': 'Invalid request'}, status=400)


